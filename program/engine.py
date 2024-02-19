from shutil import move
from socket import gethostbyname
import subprocess
from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfWriter, PdfReader
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, letter, landscape
from io import BytesIO
from datetime import datetime, timedelta
from imaplib import IMAP4_SSL
from email import message_from_bytes
from email.header import decode_header
from json import dump, load
from os import listdir, mkdir, path, rename, rmdir, remove
from re import search
from time import sleep
from tkinter import messagebox
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from pandas import ExcelWriter, read_excel, to_datetime
from psutil import AccessDenied, NoSuchProcess, process_iter


def check_accountig_folders():
    """Check if the accounting folders are created and translated and make changes if needed"""
    set_lang_file = File("ressources/lang/"+str(File("settings/param.json").get_dict()["lang"])+".json")
    lang_files = (File(f"ressources/lang/{file}") for file in listdir("ressources/lang"))
    accounting_set = set_lang_file.get_page_text("business_stats_page", "accounting")
    for lang_file in lang_files:
        accounting_lang_file = lang_file.get_page_text("business_stats_page", "accounting")

        if path.exists(accounting_lang_file):
            if accounting_lang_file != accounting_set:
                rename(accounting_lang_file, accounting_set)

            for year_folder in listdir(accounting_set):              
                months_lang_file = lang_file.get_page_text("business_stats_page", "months")
                months_set = set_lang_file.get_page_text("business_stats_page", "months")
                trimesters_lang_file = lang_file.get_page_text("business_stats_page", "trimesters")
                trimesters_set = set_lang_file.get_page_text("business_stats_page", "trimesters")

                if path.exists(accounting_set+"/"+year_folder+"/"+months_lang_file):
                    if months_lang_file != months_set:
                        rename(accounting_set+"/"+year_folder+"/"+months_lang_file,
                               accounting_set+"/"+year_folder+"/"+months_set)
                if path.exists(accounting_set+"/"+year_folder+"/"+trimesters_lang_file):
                    if trimesters_lang_file != trimesters_set:
                        rename(accounting_set+"/"+year_folder+"/"+trimesters_lang_file,
                               accounting_set+"/"+year_folder+"/"+trimesters_set)
            return
    mkdir(accounting_set)

def create_accounting_folder(folder_path, file_name):
    """Manage the folders for the accounting"""
    check_accountig_folders()
    lang_file = File("ressources/lang/"+str(File("settings/param.json").get_dict()["lang"])+".json")
    accounting = lang_file.get_page_text("business_stats_page", "accounting")
    file_path = path.join(f"{accounting}/{folder_path}", f"{file_name}.xlsx")
    # check if the file exists
    if path.exists(file_path):
        return True
    # complete or create the folder path
    current_path = f"{accounting}/"
    for folder in folder_path.split("/"):
        if not path.exists(f"{current_path}{folder}"):
            mkdir(f"{current_path}{folder}")
        current_path += f"{folder}/"

def print_msg_sheet():
    """Generates a page of thanks and print it"""    
    # create a new PDF canvas with A4 size
    packet = BytesIO()
    canvas_object = canvas.Canvas(packet, pagesize=A4)
    # get the message to print from the settings file
    msg = File("ressources/secretariat/msg.json").get_dict()["label"]
    # set the font size and position for the message
    canvas_object.setFont('Helvetica', 10)
    y = A4[1] - 20
    x = A4[0]/4

    # draw the message 12 times
    for i in range(2):
        for j in range(6):
            for line in msg.split("\n"):
                canvas_object.drawString(x - canvas_object.stringWidth(line)/2, y, line)
                y -= 15
            # padding of 20px between messages
            y = y - 20
        # reset the column
        y = A4[1] - 20
        x = A4[0]*3/4

    # save the PDF to a file
    canvas_object.save()
    pdf_writer = PdfWriter()
    pdf_reader = PdfReader(packet)
    for page in range(len(pdf_reader.pages)):
        pdf_writer.add_page(pdf_reader.pages[page])
    file_path = "labels/sheet.pdf"                  
    with open(file_path, "wb") as f:
        pdf_writer.write(f)
    
    # --- Print the sheet --- #
    try:
        print_pdf(file_path)
    except:
        lang_file = File("ressources/lang/"+str(File("settings/param.json").get_dict()["lang"])+".json")
        messagebox.showerror(lang_file.get_pop_up_text("no_pdf_viewer_title"), lang_file.get_pop_up_text("no_pdf_viewer_desc"))

    # deletes the pdf
    remove(file_path)

def print_pdf(file_path: str):
    """Prints the given pdf file"""
    try:
        # Open the PDF file with the default associated application on Windows
        subprocess.run(["start", file_path], shell=True)
        
        # max timeout is 30 seconds
        timeout = 30
        waited_time = 0
        adobe_process_running = True

        while adobe_process_running:
            # Checks if Adobe Acrobat Reader has been launched and if the timeout has not been reached
            if any(process.name() == "AcroCEF.exe" for process in process_iter()) or waited_time >= timeout:
                # Close Adobe Acrobat Reader using taskkill
                subprocess.run(["taskkill", "/f", "/im", "Acrobat.exe"], shell=True)
                break
            sleep(1)
            waited_time += 1

    except subprocess.CalledProcessError as e:
        print(f"Error opening the PDF file: {e}")
    except AccessDenied as e:
        print(f"Access denied: {e}")
    except NoSuchProcess as e:
        print(f"No such process: {e}")

def is_internet_connected():
    """Returns True if the machine is connected to the Internet"""
    try:
        # Attempt to resolve a domain name to an IP address
        gethostbyname("www.google.com")
        return True
    except:
        return False

class Stack:
    """Generates a stack to create an history of visited pages"""

    def __init__(self):
        self.pages_stack = []
        
    def unstack(self):
        """Returns the last visited page"""
        return self.pages_stack.pop()
    
    def stack(self, page):
        """Add a page to the history"""
        self.pages_stack.append(page)

    def is_empty(self) -> bool:
        """Returns True if the stack is empty"""
        return len(self.pages_stack) == 0

class File:
    """Access to text and translation for Logistic"""

    def __init__(self, path:str):
        self.path = path

    def read_data(self):
        return read_excel(self.path, engine='openpyxl')

    def sort_by_date(self):
        # get the DataFrame
        df = self.read_data()
        # Convert the "Date" column to datetime
        df['Date'] = to_datetime(df['Date'], dayfirst=True)
        # Sort the data by the Date column (newest to oldest)
        df = df.sort_values(by='Date', ascending=False)
        # Convert the "Date" column back to the desired format (e.g., "dd/mm/yyyy")
        df['Date'] = df['Date'].dt.strftime('%d/%m/%Y')

        # Save the sorted data back to the Excel file
        writer = ExcelWriter(self.path, engine='openpyxl')
        df.to_excel(writer, index=False)

        # Get the openpyxl workbook and worksheet objects
        worksheet = writer.sheets['Sheet1']

        # Adjust the width of the columns
        for column in df.columns:
            max_col_width = max(
                len(str(column)),  # Width of the column name
                df[column].apply(lambda x: len(str(x))).max(),  # Maximum width of values
            )
            col_index = df.columns.get_loc(column) + 1
            col_letter = get_column_letter(col_index)
            worksheet.column_dimensions[col_letter].width = max_col_width + 2

        # Save the changes to the Excel file
        writer._save()

    def set_pref(self, key, value):
        """Changes the selected language
        
        Argument:
            - {string} key  : preference
            - {string} value: value of the preference 
        """
        param = File("settings/param.json")
        content = param.get_dict()        
        content[key] = value    
        with open(param.path, "w", encoding="utf8") as f:
            dump(content, f, indent=4)
    
    def set_stock(self, key, value):
        """Updates the stock of a product
        
        Argument:
            - {string} key
            - value
        """
        content = self.get_dict()        
        content[key] = value    
        with open(self.path, "w") as f:
            dump(content, f, indent=4)

    def get_dict(self) -> dict:
        """Returns the content of the file"""
        with open(self.path, encoding="utf8") as f:
            return load(f)
    
    def get_page_text(self, page:str, element:str) -> str:
        """Returns the text of the element of the page"""
        return self.get_dict()["gui"][page][element]
    
    def get_toolbar_text(self, sub_menu:str, element:str) -> str:
        """Returns the text of the element of the sub menu (sub_menu or tooltips)"""
        return self.get_dict()["gui"]["toolbar"][sub_menu][element]

    def get_pop_up_text(self, element:str) -> str:
        """Returns the text associated to the element"""
        return self.get_dict()["pop_ups"][element]

    def getMenuValue(self, menu, sub=None, key="") -> str:
        """Returns the word associated with the key for a menu
        
        Args:
            - {str} menu: menu on the toolbar
            - {str} sub : sub-menu on the toolbar
            - {str} key : option name
        """
        lang = File("settings/param.json").get_dict()["lang"]
        if sub:
            return self.get_dict()[lang][menu][sub][key]
        return self.get_dict()[lang][menu][key]

    def insertLine(self, line: dict, fieldnames: list[str]):
        """Insert a new line with the given data into an XLSX file"""
        # Load the workbook from the file
        if path.exists(self.path):
            workbook = load_workbook(self.path)
            sheet = workbook.active
            # Write the headers only if the sheet is empty
            if sheet.max_row == 1 and not sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                sheet.append(fieldnames)
        else:
            workbook = Workbook()        
            sheet = workbook.active
            # Write the headers only if the sheet is empty
            if sheet.max_row == 1:
                sheet.append(fieldnames)
        # Write the values
        values = [line.get(fieldname, "") for fieldname in fieldnames]
        sheet.append(values)
        # Save the workbook with the appended data
        workbook.save(self.path)

class Email:
    """Accessor and email manager"""

    def __init__(self, email: str, password: str, primary):
        """Email account defined by username and password
        
        Args:
            - {str} email    : email of the user
            - {str} password : password of the user
            - primary        : main window for gui
        """
        # --- Checks if the user is connected to the required apps --- #
        if not primary.is_connected():
            messagebox.showwarning(
                    primary.lang_file.get_pop_up_text("vinted_connection_error_title"),
                    primary.lang_file.get_pop_up_text("vinted_connection_error_desc")
            )
            return
        self.email = email
        self.password = password
        self.param = File("settings/param.json")
        self.emails_templates_f = File("ressources/secretariat/emails_templates.json")
        self.primary = primary
        self.vinted_api = self.primary.vinted_api
        self.login()

    def login(self):
        """Connection to the email account"""
        self.imap = IMAP4_SSL("imap.gmail.com")
        try:
            self.imap.login(self.email, self.password)
        except:
            lang_file = File("ressources/lang/"+str(File("settings/param.json").get_dict()["lang"])+".json")
            messagebox.showerror(lang_file.get_pop_up_text("inc_pswd_title"), lang_file.get_pop_up_text("inc_pswd_desc"))

    def make_directories(self):
        """Folders needed to sort emails"""
        self.ROOT = "Logistic"
        BACKUP = "backup"
        SHIPPERS = "shippers"
        self.APP_LANG = self.param.get_dict()["lang"]

        # --- Check for the root --- #
        mailbox_list = self.imap.list()[1]
        exist = False
        i = 0
        while i <= len(mailbox_list):
            if self.ROOT in mailbox_list[i].decode():
                exist = True
                break
            i += 1
        if not exist:
            self.imap.create(self.ROOT)        
        
        # --- Check for the 2 branches --- #
        for lang in self.emails_templates_f.get_dict()["directories"].keys():
            dir = self.emails_templates_f.get_dict()["directories"][lang]
            for key in [BACKUP, SHIPPERS]:
                val = dir[key]
                full_folder_path = f"{self.ROOT}/{val}"
                # if the folder exists but the language is not up to date
                app_dir_value = self.emails_templates_f.get_dict()["directories"][self.APP_LANG][key]
                if app_dir_value != val:
                    # name of the directory if it exists
                    content_dir_temp = self.imap.list(full_folder_path)[1][0]
                    if not content_dir_temp == None:
                        if val in "".join(content_dir_temp.decode()):
                            self.imap.rename(full_folder_path, f"{self.ROOT}/{app_dir_value}")
        # --- Check if the main folders exist --- #
        for key in [BACKUP, SHIPPERS]:
            app_dir_value = self.emails_templates_f.get_dict()["directories"][self.APP_LANG][key]
            full_folder_path = f"{self.ROOT}/{app_dir_value}"
            # check if the folder exists
            if self.imap.list(full_folder_path)[1][0] == None:
                self.imap.create(f"{self.ROOT}/{app_dir_value}")
        # --- Check for the sub folders of shippers dir --- #
        folders = [
            "MondialRelay",
            "Chronopost",
            "Colissimo",
            "ColisPrive",
            "RelaisColis"
        ]
        shippers_dir = self.emails_templates_f.get_dict()["directories"][self.APP_LANG][SHIPPERS]
        shippers_abs_path = f"{self.ROOT}/{shippers_dir}"
        sub_mailbox_list = self.imap.list(shippers_abs_path)[1]

        for folder in folders:
            exist = False
            i = 0
            while i < len(sub_mailbox_list):
                if folder in sub_mailbox_list[i].decode():
                    exist = True
                    break
                i += 1
            if not exist:
                self.imap.create(f"{shippers_abs_path}/{folder}")
        
    def delete(self, key: str, email_id: bytes):
        """Delete an e-mail list

        Args:
            - {str} key: key to access to parameter
            - {bytes} email_id: id of the email
        """
        if self.param.get_dict()[key] == False:
            backup_dir_name = self.emails_templates_f.get_dict()["directories"][self.param.get_dict()["lang"]]["backup"]
            self.imap.copy(email_id, f'Logistic/{backup_dir_name}')
        self.imap.store(email_id, '+FLAGS', '\\Deleted')
        self.imap.expunge()

    def getMails(self):
        """Get all e-mails"""
        self.imap.select("inbox")
        # search for all emails in the inbox and return their IDs
        messages = self.imap.search(None, "ALL")[1]
        # convert the list of IDs into a list of email message IDs
        if messages[0].split(b' ') == [b'']:
            return []
        return messages[0].split(b' ')
            
    def get_codes(self):
        """Get Mondial Relay pickup codes"""
        if self.subject == "Votre colis vous attend dans votre Point Relais" and self.sender == "Mondial Relay <noreply@mondialrelay.fr>":
            # removable code
            code = self.search_pattern("Code de retrait", "Pensez")
            # pickup point
            pickup_point = self.search_pattern("Votre Point Relais®", "Merci")
            # market place
            try:
                tracking_number = self.search_pattern("colis Leboncoin", "est")
            except:
                tracking_number = self.search_pattern("colis VINTED", "est")

            # Parse the date string into a datetime object
            package = {
                "Point relais": pickup_point,
                "Numéro de suivi": tracking_number,
                "Arrivé le": self.date.strftime("%A, %d %B %Y %H:%M"),
                "Repart le": (self.date + timedelta(days=8)).strftime("%A, %d %B %Y %H:%M"),
                "Code de retrait": code
            }                        
            File("accounting/colis_mondial_relay.xlsx").insertLine(package, ["Point relais", "Numéro de suivi", "Arrivé le", "Repart le", "Code de retrait"])
            self.delete("delete-notifs", self.email_id)
            return True

    def accounting(self):
        """Update the accounting file"""
        lang_file = File("ressources/lang/"+str(File("settings/param.json").get_dict()["lang"])+".json")
        json_data = self.emails_templates_f.get_dict()["accounting"]
        list_plateforms = list(json_data.keys())
        for plateform in list_plateforms:
                bound1 = json_data[plateform]["bound1"]
                bound2 = json_data[plateform]["bound2"]

                if bound1 in self.subject and bound2 in self.subject:
                    # checking for the month file   
                    accounting = lang_file.get_page_text("business_stats_page", "accounting")
                    months = lang_file.get_page_text("business_stats_page", "months")
                    year = self.date.year
                    month = self.date.month
                    create_accounting_folder(f"{year}/{months}", month)
                    
                    # --- LeBonCoin purchase --- #
                    if "no.reply@leboncoin.fr" in self.sender:
                        if plateform == "leboncoin_purchase":
                            product = search(rf"{bound1}\s*(.*?)\s*{bound2}", self.subject).group(1).strip()
                            price = self.search_pattern("payé", "€").replace(",", ".")
                            transaction = {
                                "Date": self.date.strftime("%d/%m/%y"),
                                "Transaction": f"-{price}",
                                "Article": product,
                                "Refund": False
                            }
                        elif plateform == "leboncoin_cancel":
                            product = search(rf"{bound1}s*(.*?)\s*{bound2}", self.subject).group(1).strip()
                            price = self.search_pattern("Total remboursé", "€").replace(",", ".")
                            transaction = {
                                "Date": self.date.strftime("%d/%m/%y"),
                                "Transaction": f"{price}",
                                "Article": product,
                                "Refund": True
                            }                                    
                    # --- AliExpress purchase --- #
                    elif "<transaction@notice.aliexpress.com>" in self.sender:
                        data = self.search_pattern("Store", "Order ID")
                        total_price = self.search_pattern("total", "€Check").replace(",", ".")
                        items = []
                        def extract_item_infos(data:str):
                            """Extracts informations about the purchase"""
                            index_x = data.find("x", 0)    
                            sub_seq_end = index_x + 1

                            while sub_seq_end < len(data) and data[sub_seq_end].isdigit():
                                sub_seq_end += 1
                            # extract an item from the string
                            sub_seq = data[0:sub_seq_end]
                            items.append(sub_seq)
                            
                            # stop condition
                            if sub_seq_end == len(data):
                                return
                            return extract_item_infos(data[sub_seq_end:])
                        extract_item_infos(data)
                        items = " + ".join(items)                        
                        transaction = {
                            "Date": self.date.strftime("%d/%m/%y"),
                            "Transaction": f"-{total_price}",
                            "Article": items,
                            "Refund": False
                        }
                    # URSSAF payment
                    elif "do_not_reply@payzen.eu" in self.sender and "URSSAF" in self.sender:
                        price = self.search_pattern("votre demande de paiement de", "EUR").replace(",", ".")
                        transaction = {
                            "Date": self.date.strftime("%d/%m/%y"),
                            "Transaction": f"-{price}",
                            "Article": "Impôts URSSAF",
                            "Refund": False
                        }                      
                        
                    # save the transaction
                    File(f"{accounting}/{year}/{months}/temp-{month}.xlsx").insertLine(transaction, ["Date", "Transaction", "Article", "Refund"])
                    self.delete("delete-bills", self.email_id)
                    return True
        return
    
    def search_pattern(self, bound1: str, bound2: str) -> str:
        """Returns a pattern inside an email between two boundaries
        
        Args:
            - {str} bound1: first boundary 
            - {str} bound2: second boundary
        """
        for part in self.email_message.walk():
            if part.get_content_type() == "text/html":
                match = search(rf"{bound1}\s*(.*?)\s*{bound2}", BeautifulSoup(part.get_payload(decode=True), "html.parser").get_text(strip=True))
                if match:
                    return match.group(1).strip()

    def shipping_labels(self):
        """Process the emails link to shipping labels"""
        if self.sender in ["Team Vinted <no-reply@vinted.fr>", '"Team vinted.fr" <no-reply@vinted.fr>']:
            # --- Shipping label --- #
            if self.subject == "You’ve sold an item on Vinted":
                # check if it is a template
                product = self.search_pattern("has bought", "via")
                if product in listdir("products/templates/"):
                    # --- Get the infos from the database --- #
                    f = File(f"products/templates/{product}/infos.json")
                    infos_item = f.get_dict()
                    nb_products = infos_item["nb_products"]
                    # prevent the user if the stock is low
                    # TODO 
                    # create a custom window to refill the stocks easily when a prompt appears
                    if 1 <= nb_products <= 5:
                        messagebox.showwarning("Low Stock", f"You have only {nb_products} {product} left")
                        if nb_products == 0:
                            messagebox.showwarning("Out of Stock", f"You do not have {product} anymore.\nMake sure to refill it")
                    if not nb_products == 0:
                        # reupload the product
                        self.vinted_api.upload_item(infos_item["item"], f"products/templates/{product}/photos/", False)
                    # update the stock
                    f.set_stock("nb_products", nb_products-1)
                    
                # --- Generates shipping label --- #
                if self.get_shipping_labels() is True:
                    return True
                
            if "shipping label" in self.subject:
                # downloads shipping labels
                if self.download_shipping_labels() is True:
                    return True

    def get_shipping_labels(self):
        """Generates shipping label"""
        for part in self.email_message.walk():
            # --- Message thread --- #
            if part.get_content_type() == "text/html":
                soup = BeautifulSoup(part.get_payload(decode=True), "html.parser")
                for link in soup.find_all('a', href=True):
                    # --- Get the thread url --- #
                    try:
                        # --- Vinted --- #
                        link['href'].index("https://www.vinted.fr/e/conversation?id=")
                        # url found
                        thread = link['href']
                        self.vinted_api.get_shipping_label(thread)
                        self.delete("delete-notifs", self.email_id)
                        return True
                    except:
                        pass

    def create_shipping_label(self, shipping_label, product, company, orientation):
        """Creates the shipping label by merging the official shipping label with a custom message"""
        # --- Accessors --- #
        pdf_reader = PdfReader(shipping_label)
        pdf_writer = PdfWriter()
        
        for page_num in range(len(pdf_reader.pages)):
            # page to be modified
            page = pdf_reader.pages[page_num]
            packet = BytesIO()
            canvas_object = canvas.Canvas(packet, pagesize=orientation)
            canvas_object.setFont('Helvetica', 10)
                    
            # --- Mondial Relay and Ups --- #
            if company in ["mondialrelay", "ups"]:
                msg = f"{product}\n\n" + File("ressources/secretariat/msg.json").get_dict()["label"]
                if company == "mondialrelay":
                    y = pdf_reader.pages[0].mediabox.height/2 - 0.11*pdf_reader.pages[0].mediabox.height
                else:  # company == "ups"
                    y = pdf_reader.pages[0].mediabox.height - 0.11*pdf_reader.pages[0].mediabox.height

                for line in msg.split("\n"):
                    x = pdf_reader.pages[0].mediabox.width / (4 if company == "mondialrelay" else (3 / 2))
                    x -= canvas_object.stringWidth(line)/2
                    canvas_object.drawString(x, y, line)
                    y -= 0.018*pdf_reader.pages[0].mediabox.height
                    
            else:
                # --- Colissimo --- #
                if company == "La Poste":
                    x, y = 1/10*int(pdf_reader.pages[0].mediabox.width), 1/5*int(pdf_reader.pages[0].mediabox.height) - 0.07*int(pdf_reader.pages[0].mediabox.height)

                # --- Chronopost --- #
                if company == "chronopost":
                    x, y = 0.035*pdf_reader.pages[0].mediabox.width, 0.034*pdf_reader.pages[0].mediabox.height

                # --- Relais Colis --- #
                if company == "relaiscolis":
                    x, y = 3/4*pdf_reader.pages[0].mediabox.width - canvas_object.stringWidth(product)/2, 2/3*pdf_reader.pages[0].mediabox.height - 0.034*pdf_reader.pages[0].mediabox.height
                        
                # --- Message --- #
                canvas_object.drawString(x, y, product)
                        
        # --- Apply the modifications on the PDF --- #
        canvas_object.save()
        # move to the beginning of the StringIO buffer
        packet.seek(0)
        new_pdf = PdfReader(packet)
        # merge the modified PDF page with the original
        page.merge_page(new_pdf.pages[0])
        pdf_writer.add_page(page)

        # --- Save the shipping label --- #
        n = 1
        p = product
        while f"{product}.pdf" in listdir("labels"):
            # update of file name if it is not alone
            product = p + "("+str(n)+")"
            n += 1

        with open(f"labels/{product}.pdf", "wb") as f:
            pdf_writer.write(f)

        # --- Print the shipping label --- #
        try:
            print_pdf(f"labels/{product}.pdf")
        except:
            lang_file = File("ressources/lang/"+str(File("settings/param.json").get_dict()["lang"])+".json")
            messagebox.showerror(lang_file.get_pop_up_text("no_pdf_viewer_title"), lang_file.get_pop_up_text("no_pdf_viewer_desc"))

        # --- Backup the shipping label --- #
        if self.param.get_dict()["del-ship-label"]:
            rmdir(f"labels/{product}.pdf")
        else:
            if not path.exists("labels/backup"):
                mkdir("labels/backup")
            # rename the file if needed
            n = 1
            product = p
            while f"{product}.pdf" in listdir(f"labels/backup"):
                # update of file name if it is not alone
                product = p + "("+str(n)+")"
                n += 1
            move(f"labels/{p}.pdf", f"labels/backup/{product}.pdf")
        
        # --- Process the email --- #
        self.delete("delete-notifs", self.email_id)
      
    def download_shipping_labels(self):
        """Download shipping labels from Vinted"""
        for part in self.email_message.walk():
            # --- Vinted Attachment --- #
            if part.get_content_disposition() == "attachment":
                company = self.search_pattern("https://www\.", "\.")
                if company == None:
                    company = self.search_pattern("at a", "post office") # colissimo
                # name of the product sold
                product = self.subject[:self.subject.index("shipping label")].strip()
                attachment_bytes = part.get_payload(decode=True)
                self.create_shipping_label(BytesIO(attachment_bytes), product, company, landscape(letter))
        return True
    
    def delete_notifications(self):
        """Delete all notification emails"""
        if self.sort_notifications_emails()==True or self.sort_shipping_emails()==True:
            return True

    def folder_exists(self, folder_path):
        folder_list = self.imap.list()[1]

        for folder in folder_list:
            # Extract the folder name from the response
            folder_name = folder.decode().partition(' "/" ')[2]
            # Check if the folder path matches
            if folder_path == folder_name.strip('"'):
                return True
        return False

    def sort_notifications_emails(self):
        json_data = self.emails_templates_f.get_dict()["notifications_templates"]
        list_plateforms = list(json_data.keys())
        for plateform in list_plateforms:
            emails = json_data[plateform]["email"]
            path = json_data[plateform]["path"]
            notifications = json_data[plateform]["notifications"]
            for email in emails:
                if email in self.sender:
                    # Check if folder exists
                    try:
                        if not self.folder_exists(path):
                            self.imap.create(path)
                    except:
                        lang_file = File("ressources/lang/"+str(File("settings/param.json").get_dict()["lang"])+".json")
                        no_int_con_desc1 = lang_file.get_pop_up_text("no_int_con_desc1")
                        no_int_con_desc2 = lang_file.get_pop_up_text("no_int_con_desc2")
                        messagebox.showerror(lang_file.get_pop_up_text("no_int_con_title"), f"{no_int_con_desc1}{path}{no_int_con_desc2}")
                        return
                    # --- Template of the email is known --- #
                    for notif in notifications:
                        try:
                            if search(rf"{notif[0]}\s*(.*?)\s*{notif[1]}", self.subject):
                                self.delete("delete-notifs", self.email_id)
                                return True
                        except:
                            pass
                    # --- Template of the email is unknown --- #
                    # make a copy of the email
                    self.imap.copy(self.email_id, path)
                    self.delete("delete-notifs", self.email_id)
                    return True
                    
    def sort_shipping_emails(self):
        """Sort the emails from shippers inside different folders"""
        shippers_email = self.emails_templates_f.get_dict()["shippers_email"]
        for shipper in shippers_email.keys():
            for email_adress in shippers_email[shipper]["emails"]:
                if email_adress in self.sender:
                    try:
                        shippers_dir = self.emails_templates_f.get_dict()["directories"][self.APP_LANG]["shippers"]
                        self.imap.copy(self.email_id, f'{self.ROOT}/{shippers_dir}/{shippers_email[shipper]["name"]}')
                        self.delete("delete-notifs", self.email_id)
                        return True
                    except:
                        return
                
    def parseEmail(self, email_id: bytes):
        """Returns the email as a Python object"""
        try:
            return message_from_bytes(self.imap.fetch(email_id, '(RFC822)')[1][0][1])
        except:
            pass

    def parseAttribute(self, attribute: str):
        """Returns the parsed attribute of the email"""
        try:
            for part in decode_header(self.email_message[attribute]):
                decoded_part = part[0]
                if isinstance(decoded_part, bytes):
                    parsed_value = decoded_part.decode('utf-8')
                else:
                    parsed_value = decoded_part
            if attribute == "Date":
                if parsed_value[:22][-1] == ":":
                    parsed_value = parsed_value[:21]
                else:
                    parsed_value = parsed_value[:22]
                # Parse the date string into a datetime object
                parsed_value = datetime.strptime(parsed_value, "%a, %d %b %Y %H:%M")

            return parsed_value
        except:
            return

    def refresh_emails(self):
        """Process the emails"""
        self.make_directories()

        i = 0
        while i < len(self.getMails()):
            self.email_id = self.getMails()[i]
            # parse the email message into a Python object
            self.email_message = self.parseEmail(self.email_id)            
            try:
                # --- Attributes of the email --- #
                self.date = self.parseAttribute("Date")
                subject = self.parseAttribute("Subject")
                self.subject = ' '.join(subject.strip().split())
                self.sender = self.parseAttribute("From")
                print(self.sender)
                if self.shipping_labels() is not True:
                    if self.accounting() is not True:                        
                        if self.get_codes() is not True:
                            if self.delete_notifications() is not True:
                                # email not deleted so increase the index
                                i += 1
            except:
                # next email if it failed
                i += 1